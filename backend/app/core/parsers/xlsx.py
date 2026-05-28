import io

import openpyxl

from app.core.parsers.base import BaseParser


class XlsxParser(BaseParser):
    supported_mimes = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ]

    def parse(self, file_bytes: bytes, filename: str) -> str:
        if not file_bytes:
            raise ValueError("文件内容为空")

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        parts = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            md = self._sheet_to_markdown(sheet, sheet_name)
            if md:
                parts.append(md)

        wb.close()

        if not parts:
            raise ValueError("Excel 内容为空或无法解析")

        return "\n\n---\n\n".join(parts)

    def _sheet_to_markdown(self, sheet, sheet_name: str) -> str:
        """将单个 sheet 转为 Markdown 表格"""
        rows = []
        max_cols = 0

        for row in sheet.iter_rows(values_only=True):
            # 全为空则跳过
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            cells = []
            for cell in row:
                text = "" if cell is None else str(cell).strip()
                text = text.replace("|", "\\|").replace("\n", " ")
                cells.append(text)
            if cells:
                rows.append(cells)
                max_cols = max(max_cols, len(cells))

        if not rows:
            return ""

        # 补齐列数
        for r in rows:
            while len(r) < max_cols:
                r.append("")

        lines = [f"## {sheet_name}", ""]
        lines.append("| " + " | ".join(rows[0]) + " |")
        lines.append("| " + " | ".join(["---"] * max_cols) + " |")
        for row in rows[1:]:
            lines.append("| " + " | ".join(row) + " |")

        return "\n".join(lines)
